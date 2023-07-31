from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import csv


class ExcelEdit:
    def __init__(self):
        self.wb = None
        self.worksheet = None
        self.RowLists = []
        self.work_folder = ""

    def WriteRow(self, row_num, texts):
        for idx, text in enumerate(texts):
            self.worksheet.cell(row=row_num, column=idx + 1).value = text

    def SaveExcel(self, result_dir):
        self.wb.save(f"./{result_dir}/items.xlsm")

    def AppendRowList(self, val):
        self.RowLists.append(val)

    def LoadExcel(self, file_path):
        self.wb = load_workbook(
            file_path,
            keep_vba=True)

        self.worksheet = self.wb['items']
        item_data = []
        for row in self.worksheet.rows:
            data_row = []
            for cell in row:
                data_row.append(cell.value)
            item_data.append(data_row)

        self.worksheet = self.wb['folders']
        folder_data = []
        for row in self.worksheet.rows:
            data_row = []
            for col, cell in enumerate(row):
                if col != 0:
                    folder_data.append(cell.value)

        return(folder_data, item_data)
